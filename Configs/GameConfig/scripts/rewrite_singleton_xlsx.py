#!/usr/bin/env python3
from pathlib import Path
import openpyxl

data_dir = Path(r"F:\Project\BinGames\TEngine\Configs\GameConfig\Datas\fp")


def write_map(path: Path, fields, row):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "##var")
    for i, (name, typ, comment) in enumerate(fields, start=2):
        ws.cell(1, i, name)
    ws.cell(2, 1, "##type")
    for i, (name, typ, comment) in enumerate(fields, start=2):
        ws.cell(2, i, typ)
    ws.cell(3, 1, "##")
    for i, (name, typ, comment) in enumerate(fields, start=2):
        ws.cell(3, i, comment)
    for i, (name, typ, comment) in enumerate(fields, start=2):
        ws.cell(4, i, row[name])
    wb.save(path)
    wb.close()
    print("wrote", path.name)


write_map(
    data_dir / "#Global-FP全局公式与门槛.xlsx",
    [
        ("id", "int", "主键"),
        ("engulfRatio", "float", "吞噬阈值k"),
        ("volumeGainRatio", "float", "体积增长比"),
        ("speedVolumePenalty", "float", "体积移速惩罚"),
        ("speedFloorRatio", "float", "移速下限比"),
        ("contactDamageInterval", "float", "接触伤害间隔"),
        ("evoPointThreshold", "int", "进化点门槛"),
        ("waveInterval", "float", "波次间隔秒"),
        ("waveCount", "int", "波次数"),
        ("enemyBaseAggroRange", "float", "敌人基础察觉范围"),
        ("zapDamage", "float", "放电伤害"),
        ("zapRange", "float", "放电射程"),
        ("zapCooldown", "float", "放电冷却"),
    ],
    {
        "id": 1,
        "engulfRatio": 1.15,
        "volumeGainRatio": 0.30,
        "speedVolumePenalty": 0.18,
        "speedFloorRatio": 0.55,
        "contactDamageInterval": 0.8,
        "evoPointThreshold": 100,
        "waveInterval": 270,
        "waveCount": 3,
        "enemyBaseAggroRange": 12,
        "zapDamage": 25,
        "zapRange": 3.5,
        "zapCooldown": 1.2,
    },
)

write_map(
    data_dir / "#CellArena-细胞阶段场地与时间轴.xlsx",
    [
        ("id", "int", "主键"),
        ("arenaHalfSize", "float", "半边长"),
        ("foodConcurrent", "int", "同时食物数"),
        ("foodRespawnDelay", "float", "食物补充间隔"),
        ("foodBRatio", "float", "B型占比"),
        ("threatBaseCount", "int", "首波威胁数"),
        ("hazardCount", "int", "危险区数量"),
        ("hazardRadius", "float", "危险区半径"),
        ("hazardDamagePerSecond", "float", "危险区DPS"),
        ("microChoiceTime", "float", "微选择触发时间"),
        ("forceEvolveTime", "float", "强制进化时间"),
    ],
    {
        "id": 1,
        "arenaHalfSize": 28,
        "foodConcurrent": 6,
        "foodRespawnDelay": 4.5,
        "foodBRatio": 0.3,
        "threatBaseCount": 5,
        "hazardCount": 4,
        "hazardRadius": 3.5,
        "hazardDamagePerSecond": 6,
        "microChoiceTime": 270,
        "forceEvolveTime": 810,
    },
)

write_map(
    data_dir / "#CreatureArena-生物阶段场地与时间轴.xlsx",
    [
        ("id", "int", "主键"),
        ("exploreEnd", "float", "探索阶段结束"),
        ("pressureEnd", "float", "压力阶段结束"),
        ("arenaHalfSize", "float", "半边长"),
        ("enemyRespawnDelay", "float", "刷怪间隔"),
        ("enemyCapPhase1", "int", "阶段1敌人数上限"),
        ("enemyCapPhase2", "int", "阶段2敌人数上限"),
    ],
    {
        "id": 1,
        "exploreEnd": 240,
        "pressureEnd": 480,
        "arenaHalfSize": 22,
        "enemyRespawnDelay": 6,
        "enemyCapPhase1": 5,
        "enemyCapPhase2": 9,
    },
)
