#!/usr/bin/env python3
"""Remove incorrect ##group rows so Luban client export includes all fields."""
from pathlib import Path
import openpyxl

fp = Path(r"F:\Project\BinGames\TEngine\Configs\GameConfig\Datas\fp")
for path in sorted(fp.glob("#*.xlsx")):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    removed = 0
    # Find ##group rows and delete them
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, 1).value == "##group":
            ws.delete_rows(r, 1)
            removed += 1
    wb.save(path)
    wb.close()
    print(f"{path.name}: removed {removed} ##group row(s)")
